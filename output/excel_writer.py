
from openpyxl import Workbook

def write_excel(results, out_path='combined.xlsx', summary=None):
    wb = Workbook()
    wb.remove(wb.active)

    for res in results:
        ws = wb.create_sheet(res.get('origin', 'UNKNOWN'))
        ws.append(['components', 'licenses'])
        ws.append([str(res['components']), str(res['licenses'])])

    if summary:
        ws = wb.create_sheet('Summary')
        ws.append(['Tool', 'Status', 'Info'])
        for row in summary:
            ws.append(row)

    wb.save(out_path)
